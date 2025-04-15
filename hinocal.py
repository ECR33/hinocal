import datetime
import os.path
import argparse
from icecream import ic
import openpyxl
from openpyxl.styles import Alignment
import uuid

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# If modifying these scopes, delete the file token.json.
# SCOPES = ["https://www.googleapis.com/auth/calendar.readonly"]
SCOPES = ["https://www.googleapis.com/auth/calendar"]
CAL_ID = "bg48q2kl7vsojfsg5tssgui52s@group.calendar.google.com"  # 日野学園 年間行事カレンダー


def sign_in():
    """Googleにサインイン"""
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(
                    "ERROR: おそらくtokenの有効期限が切れています。-reオプションを指定して再度実行してください。"
                )
                ic(str(e))
                exit(1)
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open("token.json", "w") as token:
            token.write(creds.to_json())
    return creds


def update_event(service, event):
    """カレンダーにイベントを登録する"""
    result = None
    g_event = None
    try:
        g_event = (
            service.events().get(calendarId=CAL_ID, eventId=f"{event['id']}").execute()
        )
    except HttpError as e:
        # probably, not found. Let's insert new event.
        result = service.events().insert(calendarId=CAL_ID, body=event).execute()
    if (
        g_event != None
        and g_event["start"]["dateTime"] == event["start"]["dateTime"]
        and g_event["end"]["dateTime"] == event["end"]["dateTime"]
        and g_event["summary"] == event["summary"]
        and g_event["status"] == "confirmed"
        # and g_event["description"] == event["description"]
    ):
        # NOP
        # print(f"Skipped. {g_event["summary"]}")
        pass
    else:
        # Let's update!
        if g_event:
            g_event["start"]["dateTime"] = event["start"]["dateTime"]
            g_event["end"]["dateTime"] = event["end"]["dateTime"]
            g_event["summary"] = event["summary"]
            g_event["status"] = "confirmed"
            g_event["description"] = (
                g_event["description"] + "\n" + event["description"]
            )
            result = (
                service.events()
                .update(calendarId=CAL_ID, eventId=g_event["id"], body=g_event)
                .execute()
            )
        # debug
        # ic(g_event, event)
        if result:
            print("Event created/updated: %s" % (result.get("htmlLink")))
        else:
            print("Why result is None?")
    return result


def get_events(service, start_date):
    """カレンダーからイベントを取得する"""
    try:
        if start_date:
            ym = start_date.split("-")
            start = datetime.datetime(
                year=int(ym[0]),
                month=int(ym[1]),
                day=1,
                tzinfo=datetime.timezone(datetime.timedelta(hours=9)),
            )
        else:
            now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
            start = datetime.datetime(
                year=now.year,
                month=now.month,
                day=1,
                tzinfo=datetime.timezone(datetime.timedelta(hours=9)),
            )
        print(
            "Getting the upcoming 10 events",
            start.astimezone(datetime.timezone(datetime.timedelta(hours=+9))),
        )
        events_result = (
            service.events()
            .list(
                calendarId=CAL_ID,
                timeMin=start.isoformat(),
                maxResults=10,
                singleEvents=True,
                orderBy="startTime",
            )
            .execute()
        )
        events = events_result.get("items", [])
        return events

    except HttpError as error:
        print(f"An error occurred: {error}")


def download_events(service, school_year, out_file):
    """カレンダーから指定した年度のイベントを取得し、Excelファイルを作成する。"""

    def write_row(ws, row_num, event):
        """イベントを指定した行に書き出す"""
        id = event["id"]
        st = event["start"].get("dateTime", event["start"].get("date"))
        ed = event["end"].get("dateTime", event["end"].get("date"))
        st = datetime.datetime.fromisoformat(st).replace(tzinfo=None)
        ed = datetime.datetime.fromisoformat(ed).replace(tzinfo=None)
        delta_one = (ed - st) == datetime.timedelta(days=1)
        summary = event["summary"]
        description = event.get("description", "")
        ws.cell(row_num, 1, st)
        if st.strftime("%H:%M:%S") == "00:00:00":
            # date
            ws.cell(row_num, 1).number_format = "yyyy/mm/dd"
        else:
            # datetime
            ws.cell(row_num, 1).number_format = "yyyy/mm/dd hh:mm"
        if delta_one:
            # 1日のイベントの場合、終了日は省略
            pass
        else:
            if ed.strftime("%H:%M:%S") == "00:00:00":
                # 日付のみの終了日を記録する場合、その日を含む、という表現とする
                ed = ed - datetime.timedelta(days=1)
            ws.cell(row_num, 2, ed)
        if ed.strftime("%H:%M:%S") == "00:00:00":
            # date
            ws.cell(row_num, 2).number_format = "yyyy/mm/dd"
        else:
            # datetime
            ws.cell(row_num, 2).number_format = "yyyy/mm/dd hh:mm"

        ws.cell(row_num, 3, summary)
        ws.cell(row_num, 4, description)
        ws.cell(row_num, 5, id)

        topleft = Alignment(horizontal="left", vertical="top", wrap_text=False)
        topleft_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row_num, 1).alignment = topleft
        ws.cell(row_num, 2).alignment = topleft
        ws.cell(row_num, 3).alignment = topleft
        if description.find("\n") >= 0:
            # 折返し
            ws.cell(row_num, 4).alignment = topleft_wrap
        else:
            ws.cell(row_num, 4).alignment = topleft
        ws.cell(row_num, 5).alignment = topleft
        ic(st.isoformat(), summary)

    try:

        wb = openpyxl.Workbook()
        ws = wb.active
        counter = 0

        if school_year:
            school_year = int(school_year)
        else:
            now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))
            school_year = now.year

        if out_file:
            pass
        else:
            out_file = f"calendar_sy{school_year}.xlsx"


        st = datetime.datetime(
            year=school_year,
            month=4,
            day=1,
            tzinfo=datetime.timezone(datetime.timedelta(hours=9)),
        )
        ed = datetime.datetime(
            year=school_year + 1,
            month=3,
            day=31,
            tzinfo=datetime.timezone(datetime.timedelta(hours=9)),
        )
        page_token = None
        while True:
            events_result = (
                service.events()
                .list(
                    calendarId=CAL_ID,
                    timeMin=st.isoformat(),
                    timeMax=ed.isoformat(),
                    maxResults=1000,
                    singleEvents=True,
                    orderBy="startTime",
                    pageToken=page_token,
                )
                .execute()
            )
            events = events_result.get("items", [])
            for event in events:
                counter += 1
                write_row(ws, counter, event)
            page_token = events_result.get("nextPageToken")
            if not page_token:
                break

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 30

        wb.save(out_file)

        return True

    except Exception as e:
        ic(e)


def list_calendar(service):
    """カレンダー一覧を取得する"""
    try:
        page_token = None
        while True:
            calendar_list = service.calendarList().list(pageToken=page_token).execute()
            for calendar_list_entry in calendar_list["items"]:
                print(calendar_list_entry["summary"])
                # print(calendar_list_entry)
            page_token = calendar_list.get("nextPageToken")
            if not page_token:
                break
    except HttpError as error:
        print(f"An error occurred: {error}")


def create_event_from_row(row):
    summary = row[2].value
    if summary == None or summary == "":
        return None

    today = datetime.datetime.now(
        datetime.timezone(datetime.timedelta(hours=+9))
    ).isoformat()
    event_id = row[3].value
    if event_id == None or event_id == "" or event_id == 0:
        print("new record found.", row[2].value)
        event_id = uuid.uuid4().hex
    # print("event_id", event_id)
    start = row[0].value
    if type(start) == datetime.date:
        # ic("start date", start)
        pass
    elif type(start) == datetime.datetime:
        # ic("start datetime", start)
        pass
    else:
        ic("start other", type(start))
        return None
    end = row[1].value
    if type(end) == datetime.date:
        end = end + datetime.timedelta(days=1)
        # ic("end date", end)
        pass
    elif type(end) == datetime.datetime:
        end = end + datetime.timedelta(days=1)
        # ic("end datetime", end)
        pass
    else:
        # ic("end other", type(end))
        end = start + datetime.timedelta(days=1)
    start = start.replace(
        tzinfo=datetime.timezone(datetime.timedelta(hours=+9))
    ).isoformat()
    end = end.replace(
        tzinfo=datetime.timezone(datetime.timedelta(hours=+9))
    ).isoformat()

    event = {
        "id": f"{event_id}",
        "summary": f"{summary}",
        "description": f"==== {today} 時点",
        "start": {
            "dateTime": f"{start}",
            "timeZone": "Asia/Tokyo",
        },
        "end": {
            "dateTime": f"{end}",
            "timeZone": "Asia/Tokyo",
        },
        "reminders": {"useDefault": False},
    }

    # ic(event)
    return event


def main(command=None, in_file=None, args=None):

    creds = sign_in()
    service = build("calendar", "v3", credentials=creds)

    if command == "list":
        events = get_events(service, args.startdate)

        if not events:
            print("No upcoming events found.")
            return

        # Prints the start and name of the next 10 events
        for event in events:
            start = event["start"].get("dateTime", event["start"].get("date"))
            print(start, event["id"], event["summary"], event["description"])

    if command == "calendar":
        list_calendar(service)

    if command == "download":
        download_events(service, args.school_year, args.out_file)

    if command == "sync":
        ic(in_file)
        if in_file == None:
            print("Please specify the file name for input.")
            return False
        wb = openpyxl.load_workbook(in_file)
        ws = wb["Sheet1"]
        max_row = ws.max_row
        counter = 0
        for row in ws:
            counter += 1
            # if counter > 10:  # debug
            #     break
            if row[0].value == "日付":
                # skip
                print(f"{counter}/{max_row}: skip title row")
                pass
            else:
                g_event = create_event_from_row(row)
                if g_event:
                    event = update_event(service, g_event)
                    if event:
                        print(
                            f"{counter}/{max_row}: Updated. {event["start"]["dateTime"]} - {event["end"]["dateTime"]}: {event["summary"]} {event["description"]}"
                        )
                        # write back to excel sheet
                        ws.cell(counter, 4, g_event["id"])
                    else:
                        print(
                            f"{counter}/{max_row}: Skipped. {g_event["start"]["dateTime"]} - {g_event["end"]["dateTime"]}: {g_event["summary"]}"
                        )
                else:
                    # ignore None
                    pass
        wb.save(in_file)

        try:
            pass

            # event = update_event(service, event)
            # print ('Event created: %s' % (event.get('htmlLink')))

        except HttpError as error:
            print(f"An error occurred: {error}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Googleカレンダーにイベント(予定)を登録・更新する"
    )
    parser.add_argument(
        "command",
        choices=["list", "sync", "calendar", "download"],
        help="list: Get and print events from Google calender. sync: Sync local to Google. calendar: Get and print calenders from Google.",
    )
    parser.add_argument(
        "-re",
        "--relogin",
        action="store_true",
        help="サインイン情報をクリアしてから実行する",
    )
    parser.add_argument("-sd", "--startdate", help="開始年月 yyyy-mm")
    parser.add_argument("-f", "--file", help="行事予定一覧excelファイル")
    parser.add_argument(
        "-of",
        "--out_file",
        help="カレンダーの内容を書き出すexcelファイル名。指定しない場合は'calendar_syYYYY.xlsx'。既存のファイルは上書きされる。",
    )
    parser.add_argument("-sy", "--school_year", help="年度 yyyy")
    args = parser.parse_args()
    if args.relogin:
        if os.path.exists("./token.json"):
            os.remove("./token.json")
    main(args.command, args.file, args)
